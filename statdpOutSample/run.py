import re

import openpyxl

from statdpwrapper.postprocessing import the_zero_noise_prng, CountPP, HammingDistancePP, EntryPP, LengthPP
from statdpwrapper.postprocessing import *


pat = re.compile(r'(\[.+\]) (.+)')
def parse_attack(s):
    m = pat.match(s)
    if 'inf' in m[1]:
        event = (eval(m[1].replace('inf', 'float("inf")')),)
    else:
        event = eval(m[1])
    pp = m[2]
    if not pp.endswith(')'):
        pp += '()'
    pp = eval(pp)
    return event, pp

class Run:
    def __init__(self, name, a1, a2, s, postfix):
        self.name = name
        self.a1 = a1
        self.a2 = a2
        self.event, self.pp = parse_attack(s)
        self.a0 = a1
        self.postfix = postfix + "-out"
    def __repr__(self):
        return self.__dict__.__repr__()

MECH_MAP = {
    'LaplaceMechanism': 'LaplaceMechanism',
    'TruncatedGeometric': 'TruncatedGeometricMechanism',
    'NoisyHistogram1': 'NoisyHist1',
    'NoisyHistogram2': 'NoisyHist2',
    'OneTimeRAPPOR': 'OneTimeRappor',
    'RAPPOR': 'Rappor',
    'SVT1': 'SparseVectorTechnique1',
    'SVT2': 'SparseVectorTechnique2',
    'SVT3': 'SparseVectorTechnique3',
    'SVT4': 'SparseVectorTechnique4',
    'SVT5': 'SparseVectorTechnique5',
    'SVT6': 'SparseVectorTechnique6',
    'ReportNoisyMax1': 'ReportNoisyMax1',
    'ReportNoisyMax2': 'ReportNoisyMax2',
    'ReportNoisyMax3': 'ReportNoisyMax3',
    'ReportNoisyMax4': 'ReportNoisyMax4',
}

def get_run(row):
    algname, a1, a2, s, eps, inf_eps, p1, p2, method = (c.value for c in row)
    newalgname = MECH_MAP[algname]
    a1, a2 = eval(a1), eval(a2)
    assert isinstance(a1, list)
    assert isinstance(a2, list)
    return Run(newalgname, a1, a2, s, algname)


runs = []

def parse_xls(file):
    book = openpyxl.load_workbook(file)
    for sh in book.worksheets:
        for row in sh.iter_rows(2, sh.max_row):
            runs.append(get_run(row))

parse_xls('/media/barry/Data/DPDisprover/results/statdp/statdp_out.xlsx')

# runs = [
#     Run('SparseVectorTechnique5',
#         [2, 0, 0, 0, 0, 0, 0, 0, 0, 0],
#         [1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
#         [9],
#         CountPP(False),
#         'SVT5'
#     ),
#     Run('SparseVectorTechnique5',
#         [2, 2, 2, 2, 2, 0, 0, 0, 0, 0],
#         [1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
#         [5],
#         CountPP(True),
#         'SVT5'
#     ),
#     Run('SparseVectorTechnique5',
#         [2, 0, 0, 0, 0, 0, 0, 0, 0, 0],
#         [1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
#         [9],
#         HammingDistancePP(),
#         'SVT5'
#     ),
#     Run('SparseVectorTechnique5',
#         [0, 2, 2, 2, 2, 2, 2, 2, 2, 2],
#         [1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
#         [1],
#         CountPP(False),
#         'SVT5'
#     ),
#     Run('SparseVectorTechnique6',
#         [0, 0, 0, 0, 0, 1, 1, 1, 1, 1],
#         [1, 1, 1, 1, 1, 0, 0, 0, 0, 0],
#         [9],
#         HammingDistancePP(),
#         'SVT6'
#     ),
#     Run('SparseVectorTechnique6',
#         [1, 1, 1, 1, 1, 0, 0, 0, 0, 0],
#         [0, 0, 0, 0, 0, 1, 1, 1, 1, 1],
#         [1],
#         HammingDistancePP(),
#         'SVT6'
#     ),
#     Run('SparseVectorTechnique6',
#         [0, 0, 0, 0, 0, 1, 1, 1, 1, 1],
#         [1, 1, 1, 1, 1, 0, 0, 0, 0, 0],
#         [9],
#         HammingDistancePP(),
#         'SVT6'
#     ),
#     Run('SparseVectorTechnique1',
#         [1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
#         [2, 2, 2, 2, 2, 0, 0, 0, 0, 0],
#         [6],
#         HammingDistancePP(),
#         'SVT1'
#     ),
#     Run('SparseVectorTechnique1',
#         [1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
#         [2, 2, 2, 2, 2, 0, 0, 0, 0, 0],
#         [5],
#         CountPP(False),
#         'SVT1'
#     ),
#     Run('SparseVectorTechnique4',
#         [0, 0, 0, 0, 0, 1, 1, 1, 1, 1],
#         [1, 1, 1, 1, 1, 0, 0, 0, 0, 0],
#         [5],
#         CountPP(False),
#         'SVT4'
#     ),
#     Run('SparseVectorTechnique4',
#         [0, 0, 0, 0, 0, 1, 1, 1, 1, 1],
#         [1, 1, 1, 1, 1, 0, 0, 0, 0, 0],
#         [5],
#         HammingDistancePP(),
#         'SVT4'
#     ),
#     Run('SparseVectorTechnique4',
#         [1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
#         [2, 2, 2, 2, 2, 0, 0, 0, 0, 0],
#         [6],
#         CountPP(False),
#         'SVT4'
#     ),
#     Run('ReportNoisyMax1',
#         [1, 1, 1, 1, 1],[0, 2, 2, 2, 2],
#         [0],
#         EntryPP(0),
#         'NoisyMax1'
#     ),
#     Run('ReportNoisyMax1',
#         [2, 0, 0, 0, 0],[1, 1, 1, 1, 1],
#         [0],
#         EntryPP(0),
#         'NoisyMax1'
#     ),
#     Run('ReportNoisyMax2',
#         [2, 0, 0, 0, 0],[1, 1, 1, 1, 1],
#         [0],
#         EntryPP(0),
#         'NoisyMax2'
#     ),
#     Run('ReportNoisyMax2',
#         [1, 1, 1, 1, 1],[0, 2, 2, 2, 2],
#         [0],
#         EntryPP(0),
#         'NoisyMax2'
#     )
# ]