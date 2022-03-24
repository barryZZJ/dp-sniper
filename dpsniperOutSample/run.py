import re
import numpy as np
import os
import torch
import openpyxl

from dpsniper.classifiers.feature_transformer import FlagsFeatureTransformer
from dpsniper.attack.ml_attack import MlAttack
from dpsniper.classifiers.torch_optimizer_factory import AdamOptimizerFactory, SGDOptimizerFactory

from dpsniper.classifiers.classifier_factory import MultiLayerPerceptronFactory, LogisticRegressionFactory

from dpsniper.mechanisms.parallel import *
from dpsniper.mechanisms.noisy_hist import *
from dpsniper.mechanisms.rappor import *
from dpsniper.mechanisms.prefix_sum import *
from dpsniper.mechanisms.report_noisy_max import *
from dpsniper.mechanisms.geometric import TruncatedGeometricMechanism

def class_name(mechanism):
    return type(mechanism).__name__.split(".")[-1]

pat = re.compile(r't=([0-9\.]+), q=([0-9\.]+).+scale=([0-9\.\[\]\s]+), mean=([0-9\.\[\]\s]+)[\w\W]+file=(.+)')
def parse_attack(s):
    m = pat.match(s)
    t = float(m[1])
    q = float(m[2])
    scale = m[3].strip('[]\n')
    scale = np.fromstring(scale, dtype=np.float64, sep=' ')
    mean = m[4].strip('[]\n')
    mean = np.fromstring(mean, dtype=np.float64, sep=' ')
    file = m[5]
    return t,q,scale,mean,file

class Run:
    def __init__(self, mechanism, a1: list, a2: list, s:str, eps, p1, p2, is_mlp: bool, output_size: int, feature_transform=None):
        self.name = class_name(mechanism)
        self.series_name = 'dd_search_reg' if not is_mlp else 'dd_search_mlp'
        self.mechanism = mechanism
        self.a1 = np.array(a1)
        self.a2 = np.array(a2)
        self.a1_ = a1
        self.a2_ = a2
        self.eps = eps
        self.p1, self.p2 = p1, p2
        t, q, scale, mean, state_dict_file = parse_attack(s)

        logdir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'dpsniper/experiments/logs')
        for dir, _, filenames in os.walk(logdir):
            if dir.endswith('models') and state_dict_file in filenames:
                state_dict_file = os.path.join(dir, state_dict_file)
                break

        if is_mlp:
            factory = MultiLayerPerceptronFactory(
                in_dimensions=output_size,
                hidden_sizes=(10, 5),
                optimizer_factory=AdamOptimizerFactory(learning_rate=0.1,
                                                       step_size=500),
                feature_transform=feature_transform,
                regularization_weight=0.0001,
                epochs=10,
                label=class_name(mechanism),
                state_dict_file=state_dict_file,
                do_log=False
            )
        else:
            factory = LogisticRegressionFactory(
                feature_transform=feature_transform,
                in_dimensions=output_size,
                optimizer_factory=SGDOptimizerFactory(learning_rate=0.3, momentum=0.3, step_size=500),
                regularization_weight=0.001,
                epochs=10,
                label=class_name(mechanism),
                state_dict_file=state_dict_file,
                do_log=False
            )
        torch.set_default_dtype(torch.float64)
        classifier = factory.create()

        def _generate_data_batch(a1, a2, n):
            """
            Generates a training data batch of size 2n (n samples for each input a1 and a2).
            """
            b1 = mechanism.m(a1, n)
            b2 = mechanism.m(a2, n)
            if len(b1.shape) == 1:
                # make sure b1 and b2 have shape (n_samples, 1)
                b1 = np.atleast_2d(b1).T
                b2 = np.atleast_2d(b2).T

            # rows = samples, columns = features
            x = np.concatenate((b1, b2), axis=0)

            # 1d array of labels
            y = np.zeros(2 * n)
            y[n:] = 1

            return x, y

        x, _ = _generate_data_batch(self.a1, self.a2, 100_000)
        classifier.init_normalizer(x, mean, scale)
        self.classifier = classifier
        self.attack = MlAttack(classifier, t, q)

    def generate_out_prob_samples(self, n):
        b1 = self.mechanism.m(self.a1, n)
        b2 = self.mechanism.m(self.a2, n)
        if len(b1.shape) == 1:
            # make sure b1 and b2 have shape (n_samples, 1)
            b1 = np.atleast_2d(b1).T
            b2 = np.atleast_2d(b2).T

        # rows = samples, columns = features
        bs = np.concatenate((b1, b2), axis=0)
        probs = self.classifier.predict_probabilities(bs)
        return bs, probs

    def generate_respect_out_prob_samples(self, n):
        b1 = self.mechanism.m(self.a1, n)
        b2 = self.mechanism.m(self.a2, n)
        if len(b1.shape) == 1:
            # make sure b1 and b2 have shape (n_samples, 1)
            b1 = np.atleast_2d(b1).T
            b2 = np.atleast_2d(b2).T

        # rows = samples, columns = features
        probs1 = self.classifier.predict_probabilities(b1)
        probs2 = 1-self.classifier.predict_probabilities(b2)
        return b1, b2, probs1, probs2

MECH_MAP = {
    'LaplaceMechanism': [LaplaceMechanism(), 1],
    'TruncatedGeometric': [TruncatedGeometricMechanism(), 1],
    'NoisyHistogram1': [NoisyHist1(), 5],
    'NoisyHistogram2': [NoisyHist2(), 5],
    'SVT1': [SparseVectorTechnique1(c=1, t=0.5), 20],
    'SVT2': [SparseVectorTechnique2(c=1, t=1), 20],
    'SVT3': [SparseVectorTechnique3(c=1, t=1), 30],
    'SVT4': [SparseVectorTechnique4(c=1, t=1), 20],
    'SVT5': [SparseVectorTechnique5(c=1, t=1), 10],
    'SVT6': [SparseVectorTechnique6(c=1, t=1), 10],
    'ReportNoisyMax1-Lap': [ReportNoisyMax1(), 1],
    'ReportNoisyMax2-Exp': [ReportNoisyMax2(), 1],
    'ReportNoisyMax3-Lap': [ReportNoisyMax3(), 1],
    'ReportNoisyMax4-Exp': [ReportNoisyMax4(), 1],
    'OneTimeRAPPOR': [OneTimeRappor(), OneTimeRappor().filter_size],
    'RAPPOR': [Rappor(), Rappor().filter_size],
}
TRANS_MAP = {
    'SVT1': FlagsFeatureTransformer([-1]),
    'SVT2': FlagsFeatureTransformer([-1]),
    'SVT3': FlagsFeatureTransformer([-1000.0, -2000.0]),
    'SVT4': FlagsFeatureTransformer([-1]),
}

def get_run(row):
    algname, a1, a2, s, eps, inf_eps, p1, p2, method = (c.value for c in row)
    mech = MECH_MAP[algname][0]
    output_size = MECH_MAP[algname][1]
    a1, a2 = eval(a1), eval(a2)
    assert isinstance(a1, list)
    assert isinstance(a2, list)
    is_mlp = 'mlp' in method
    return Run(mech, a1, a2, s, eps, p1, p2, is_mlp, output_size, TRANS_MAP.get(algname))

runs = []

def parse_xls(file):
    book = openpyxl.load_workbook(file)
    for sh in book.worksheets:
        for row in sh.iter_rows(2, sh.max_row):
            runs.append(get_run(row))


parse_xls('/media/barry/Data/DPDisprover/results/dpsniper/dpsniper_out.xlsx')
